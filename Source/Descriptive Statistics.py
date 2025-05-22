import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import os
import statistics
import numpy as np
import pandas as pd
from tkinter import filedialog
import tkinter as tk
from scipy import stats
import matplotlib.pyplot as plt

# 定义语言字典
LANGUAGES = {
    'zh': {
        'title': "描述性统计",
        'select_button': "选择文件",
        'analyze_button': "分析文件",
        'file_not_found': "文件不存在，请重新选择。",
        'analysis_success': "分析完成，结果已保存到 {}\n",
        'no_save_path': "未选择保存路径，结果未保存。",
        'analysis_error': "分析文件时出错: {}",
        'images_saved': "图片已保存到 {}",
        'switch_language': "切换语言",
        'explanation': {
            "Mean": "均值是数据集中所有数值的平均值，反映了数据的集中趋势。",
            "Median": "中位数是将数据集按升序排列后位于中间位置的数值，它不受极端值的影响，能更好地反映数据的中间水平。",
            "Standard Deviation": "标准差衡量了数据相对于均值的离散程度，标准差越大，数据越分散。",
            "Minimum": "最小值是数据集中的最小数值。",
            "Maximum": "最大值是数据集中的最大数值。",
            "Range": "极差是最大值与最小值的差值，反映了数据的取值范围。",
            "First Quartile (Q1)": "第一四分位数是将数据集按升序排列后位于 25% 位置的数值，它将数据集分为前 25% 和后 75%。",
            "Third Quartile (Q3)": "第三四分位数是将数据集按升序排列后位于 75% 位置的数值，它将数据集分为前 75% 和后 25%。",
            "Interquartile Range (IQR)": "四分位距是第三四分位数与第一四分位数的差值，反映了数据中间 50% 的分布范围，不受极端值的影响。",
            "Kurtosis": "峰度衡量了数据分布的峰态，反映了数据在均值附近的集中程度和尾部的厚度。",
            "Skewness": "偏度衡量了数据分布的不对称程度，正值表示右偏，负值表示左偏。",
            "Mode": "众数是数据集中出现次数最多的数值。"
        }
    },
    'en': {
        'title': "Descriptive statistics",
        'select_button': "Select File",
        'analyze_button': "Analyze File",
        'file_not_found': "The file does not exist. Please select again.",
        'analysis_success': "Analysis completed. The results have been saved to {}\n",
        'no_save_path': "No save path selected. The results were not saved.",
        'analysis_error': "An error occurred while analyzing the file: {}",
        'images_saved': "Images have been saved to {}",
        'switch_language': "Switch Language",
        'explanation': {
            "Mean": "The mean is the average of all values in the dataset, reflecting the central tendency of the data.",
            "Median": "The median is the middle value when the dataset is arranged in ascending order. It is not affected by extreme values and better reflects the middle level of the data.",
            "Standard Deviation": "The standard deviation measures the dispersion of data relative to the mean. A larger standard deviation indicates more dispersed data.",
            "Minimum": "The minimum is the smallest value in the dataset.",
            "Maximum": "The maximum is the largest value in the dataset.",
            "Range": "The range is the difference between the maximum and minimum values, reflecting the range of the data.",
            "First Quartile (Q1)": "The first quartile is the value at the 25% position when the dataset is arranged in ascending order. It divides the dataset into the first 25% and the last 75%.",
            "Third Quartile (Q3)": "The third quartile is the value at the 75% position when the dataset is arranged in ascending order. It divides the dataset into the first 75% and the last 25%.",
            "Interquartile Range (IQR)": "The interquartile range is the difference between the third and first quartiles, reflecting the distribution range of the middle 50% of the data and is not affected by extreme values.",
            "Kurtosis": "Kurtosis measures the peakedness of the data distribution, reflecting the concentration of data around the mean and the thickness of the tails.",
            "Skewness": "Skewness measures the asymmetry of the data distribution. A positive value indicates right skewness, and a negative value indicates left skewness.",
            "Mode": "The mode is the value that appears most frequently in the dataset."
        }
    }
}

# 当前语言
current_language = 'en'

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)
        file_entry.configure(style="TEntry")  # 恢复默认样式

def analyze_file():
    global current_language
    file_path = file_entry.get()
    if file_path == "请输入待分析 Excel 文件的完整路径" or file_path == "Please enter the full path of the Excel file to be analyzed":
        file_path = ""
    if not os.path.exists(file_path):
        result_label.config(text=LANGUAGES[current_language]['file_not_found'])
        return
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        rows = sheet.max_row
        columns = sheet.max_column

        data = []
        columns_stats = ["Mean", "Median", "Standard Deviation", "Minimum", "Maximum", "Range",
                         "First Quartile (Q1)", "Third Quartile (Q3)", "Interquartile Range (IQR)",
                         "Kurtosis", "Skewness", "Mode"]
        explanations = LANGUAGES[current_language]['explanation']

        for col_idx in range(1, columns + 1):
            column_values = []
            for row_idx in range(2, rows + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, (int, float)):
                    column_values.append(cell_value)

            if column_values:
                col_name = sheet.cell(row=1, column=col_idx).value
                mean_val = statistics.mean(column_values)
                median_val = statistics.median(column_values)
                try:
                    stdev_val = statistics.stdev(column_values)
                except statistics.StatisticsError:
                    stdev_val = 0
                min_val = min(column_values)
                max_val = max(column_values)
                range_val = max_val - min_val
                q1 = np.percentile(column_values, 25)
                q3 = np.percentile(column_values, 75)
                iqr = q3 - q1
                kurtosis_val = stats.kurtosis(column_values)
                skewness_val = stats.skew(column_values)
                try:
                    mode_val = statistics.mode(column_values)
                except statistics.StatisticsError:
                    mode_val = None

                values = [mean_val, median_val, stdev_val, min_val, max_val, range_val,
                          q1, q3, iqr, kurtosis_val, skewness_val, mode_val]
                data.append([col_name] + values)

        headers = ["Column Name"] + columns_stats
        df = pd.DataFrame(data, columns=headers)

        # 添加解释说明
        explanation_df = pd.DataFrame([explanations])
        explanation_df = explanation_df.reindex(columns=columns_stats)
        explanation_df.insert(0, "Column Name", "解释说明" if current_language == 'zh' else "Explanation")

        # 合并数据和解释说明
        combined_df = pd.concat([df, explanation_df], ignore_index=True)

        # 转置数据框
        transposed_df = combined_df.set_index('Column Name').T.reset_index().rename(columns={'index': 'Column Name'})

        # 让用户选择保存路径
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            # 保存到 Excel 文件
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                transposed_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            result_msg = LANGUAGES[current_language]['analysis_success'].format(save_path)

            # 获取保存路径的目录
            save_dir = os.path.dirname(save_path)

            for col_idx in range(1, columns + 1):
                column_values = []
                for row_idx in range(2, rows + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)):
                        column_values.append(cell_value)

                if column_values:
                    col_name = sheet.cell(row=1, column=col_idx).value

                    # 生成图表
                    fig, axes = plt.subplots(2, 2, figsize=(10, 8))

                    # 频度分布图
                    axes[0, 0].hist(column_values, bins=20, edgecolor='k')
                    axes[0, 0].set_title(f'Frequency Distribution - {col_name}')
                    axes[0, 0].set_xlabel('Value')
                    axes[0, 0].set_ylabel('Frequency')

                    # 直方图
                    axes[0, 1].hist(column_values, bins=20, density=True, edgecolor='k')
                    axes[0, 1].set_title(f'Histogram - {col_name}')
                    axes[0, 1].set_xlabel('Value')
                    axes[0, 1].set_ylabel('Density')

                    # 箱线图
                    axes[1, 0].boxplot(column_values)
                    axes[1, 0].set_title(f'Box Plot - {col_name}')
                    axes[1, 0].set_ylabel('Value')

                    # 散点图（这里简单用索引作为 x 轴）
                    x = np.arange(len(column_values))
                    axes[1, 1].scatter(x, column_values)
                    axes[1, 1].set_title(f'Scatter Plot - {col_name}')
                    axes[1, 1].set_xlabel('Index')
                    axes[1, 1].set_ylabel('Value')

                    plt.tight_layout()

                    # 生成图片保存路径
                    img_name = f"{col_name}_charts.png"
                    img_path = os.path.join(save_dir, img_name)

                    # 保存图片
                    plt.savefig(img_path)
                    plt.close()

            result_msg += LANGUAGES[current_language]['images_saved'].format(save_dir)
            result_label.config(text=result_msg, wraplength=400)
        else:
            result_label.config(text=LANGUAGES[current_language]['no_save_path'])

    except Exception as e:
        result_label.config(text=LANGUAGES[current_language]['analysis_error'].format(str(e)))

def switch_language():
    global current_language
    current_language = 'en' if current_language == 'zh' else 'zh'
    root.title(LANGUAGES[current_language]['title'])
    select_button.config(text=LANGUAGES[current_language]['select_button'])
    analyze_button.config(text=LANGUAGES[current_language]['analyze_button'])
    language_label.config(text=LANGUAGES[current_language]['switch_language'])
    # 切换语言时更新提示信息
    file_entry.delete(0, tk.END)
    if current_language == 'zh':
        file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
        file_entry.configure(style="Gray.TEntry")
    else:
        file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
        file_entry.configure(style="Gray.TEntry")

def on_entry_click(event):
    """当用户点击输入框时，清除提示信息"""
    if file_entry.get() == "请输入待分析 Excel 文件的完整路径" or file_entry.get() == "Please enter the full path of the Excel file to be analyzed":
        file_entry.delete(0, tk.END)
        file_entry.configure(style="TEntry")  # 恢复默认样式

def on_focusout(event):
    """当用户离开输入框时，如果没有输入内容，显示提示信息"""
    if file_entry.get() == "":
        if current_language == 'zh':
            file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
        else:
            file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
        file_entry.configure(style="Gray.TEntry")

# 创建主窗口
root = ttk.Window(themename="flatly")
root.title(LANGUAGES[current_language]['title'])

# 获取屏幕宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 计算窗口的宽度和高度
window_width = 500
window_height = 250

# 计算窗口的 x 和 y 坐标，使其居中
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2

# 设置窗口的位置
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 创建一个框架来包含按钮和输入框
frame = ttk.Frame(root)
frame.pack(expand=True)  # 使用 expand 选项使框架在上下方向上居中

# 创建文件选择按钮
select_button = ttk.Button(frame, text=LANGUAGES[current_language]['select_button'], command=select_file, bootstyle=PRIMARY)
select_button.pack(pady=10)

# 创建自定义样式
style = ttk.Style()
style.configure("Gray.TEntry", foreground="gray")
style.configure("Gray.TLabel", foreground="gray")

# 创建文件路径输入框
file_entry = ttk.Entry(frame, width=50, style="Gray.TEntry")
if current_language == 'zh':
    file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
else:
    file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
file_entry.pack(pady=5)
file_entry.bind("<FocusIn>", on_entry_click)
file_entry.bind("<FocusOut>", on_focusout)

# 创建分析按钮
analyze_button = ttk.Button(frame, text=LANGUAGES[current_language]['analyze_button'], command=analyze_file, bootstyle=SUCCESS)
analyze_button.pack(pady=10)

# 创建语言切换标签
language_label = ttk.Label(frame, text=LANGUAGES[current_language]['switch_language'], style="Gray.TLabel", cursor="hand2")
language_label.pack(pady=10)
language_label.bind("<Button-1>", lambda event: switch_language())

# 创建结果显示标签
result_label = ttk.Label(root, text="", justify=tk.LEFT)
result_label.pack(pady=10)

# 运行主循环
root.mainloop()