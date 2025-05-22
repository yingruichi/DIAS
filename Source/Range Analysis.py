import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import os
import numpy as np
import pandas as pd
from tkinter import filedialog
import tkinter as tk
import matplotlib.pyplot as plt
import pathlib

# 定义语言字典
LANGUAGES = {
    'zh': {
        'title': "极差分析",
        'select_button': "选择文件",
        'analyze_button': "分析文件",
        'file_not_found': "文件不存在，请重新选择。",
        'analysis_success': "分析完成，结果已保存到 {}\n",
        'no_save_path': "未选择保存路径，结果未保存。",
        'analysis_error': "分析文件时出错: {}",
        'switch_language': "切换语言",
        'explanation': {
            "极差": "极差反映了数据的离散程度，在极差分析中，极差越大说明该因素对试验结果的影响越大。"
        },
        'interpretation': {
            "极差": "极差越大，表明该因素对试验结果的影响越显著。",
            "均值": "各水平下试验结果的平均值，用于比较不同水平对试验结果的影响。"
        }
    },
    'en': {
        'title': "Range Analysis",
        'select_button': "Select File",
        'analyze_button': "Analyze File",
        'file_not_found': "The file does not exist. Please select again.",
        'analysis_success': "Analysis completed. The results have been saved to {}\n",
        'no_save_path': "No save path selected. The results were not saved.",
        'analysis_error': "An error occurred while analyzing the file: {}",
        'switch_language': "Switch Language",
        'explanation': {
            "极差": "The range reflects the dispersion of the data. In range analysis, a larger range indicates that the factor has a greater influence on the test results."
        },
        'interpretation': {
            "极差": "A larger range indicates that the factor has a more significant influence on the test results.",
            "均值": "The average value of the test results at each level, used to compare the influence of different levels on the test results."
        }
    }
}

# 当前语言
current_language = 'en'


def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)
        file_entry.configure(style="TEntry")  # 恢复默认样式


def analyze_file():
    global current_language
    file_path = file_entry.get()
    if file_path == "请输入待分析 Excel 文件的完整路径" or file_path == "Please enter the full path of the Excel file to be analyzed":
        file_path = ""
    if not os.path.exists(file_path):
        result_label.config(text=LANGUAGES[current_language]['file_not_found'])
        return
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        rows = sheet.max_row
        columns = sheet.max_column

        data = []
        explanations = LANGUAGES[current_language]['explanation']
        interpretations = LANGUAGES[current_language]['interpretation']

        # 读取数据
        table = []
        for row_idx in range(1, rows + 1):
            row = []
            for col_idx in range(1, columns + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, (int, float)):
                    row.append(cell_value)
            if row:
                table.append(row)
        table = np.array(table)

        num_factors = columns - 1  # 减去结果列
        levels = np.unique(table[:, :-1], axis=0).shape[0]
        results = table[:, -1]

        # 计算各因素各水平下的均值和极差
        factor_means = []
        factor_ranges = []
        for factor in range(num_factors):
            level_means = []
            for level in np.unique(table[:, factor]):
                level_results = results[table[:, factor] == level]
                level_mean = np.mean(level_results)
                level_means.append(level_mean)
            factor_means.append(level_means)
            factor_range = np.max(level_means) - np.min(level_means)
            factor_ranges.append(factor_range)

        # 整理数据
        for i in range(num_factors):
            for j in range(levels):
                data.append([f"因素{i + 1} 水平{j + 1} 均值", factor_means[i][j]])
            data.append([f"因素{i + 1} 极差", factor_ranges[i]])

        headers = ["统计量", "值"]
        df = pd.DataFrame(data, columns=headers)

        # 添加解释说明
        explanation_df = pd.DataFrame([explanations])
        explanation_df = explanation_df.reindex(columns=["极差"])
        explanation_df.insert(0, "统计量", "解释说明" if current_language == 'zh' else "Explanation")

        # 添加分析结果解读
        interpretation_df = pd.DataFrame([interpretations])
        interpretation_df = interpretation_df.reindex(columns=["极差", "均值"])
        interpretation_df.insert(0, "统计量", "结果解读" if current_language == 'zh' else "Interpretation")

        # 合并数据、解释说明和结果解读
        combined_df = pd.concat([df, explanation_df, interpretation_df], ignore_index=True)

        # 转置数据框
        transposed_df = combined_df.set_index('统计量').T.reset_index().rename(columns={'index': '统计量'})

        # 让用户选择保存路径
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            # 保存到 Excel 文件
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                transposed_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # 生成极差分析图
            desktop_path = pathlib.Path.home() / 'Desktop'
            plot_path = desktop_path / 'range_analysis_plot.png'
            plt.figure(figsize=(10, 6))
            for i in range(num_factors):
                plt.plot(np.arange(1, levels + 1), factor_means[i], marker='o', label=f'Factor {i + 1}')
            plt.title('Range Analysis - Mean Values by Factor and Level')
            plt.xlabel('Level')
            plt.ylabel('Mean Value')
            plt.legend()
            plt.grid(True)
            plt.savefig(plot_path)
            plt.close()

            result_msg = LANGUAGES[current_language]['analysis_success'].format(
                save_path) + f"极差分析图已保存到 {plot_path}"
            result_label.config(text=result_msg, wraplength=400)
        else:
            result_label.config(text=LANGUAGES[current_language]['no_save_path'])

    except Exception as e:
        result_label.config(text=LANGUAGES[current_language]['analysis_error'].format(str(e)))


def switch_language():
    global current_language
    current_language = 'en' if current_language == 'zh' else 'zh'
    root.title(LANGUAGES[current_language]['title'])
    select_button.config(text=LANGUAGES[current_language]['select_button'])
    analyze_button.config(text=LANGUAGES[current_language]['analyze_button'])
    language_label.config(text=LANGUAGES[current_language]['switch_language'])
    # 切换语言时更新提示信息
    file_entry.delete(0, tk.END)
    if current_language == 'zh':
        file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
        file_entry.configure(style="Gray.TEntry")
    else:
        file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
        file_entry.configure(style="Gray.TEntry")


def on_entry_click(event):
    """当用户点击输入框时，清除提示信息"""
    if file_entry.get() == "请输入待分析 Excel 文件的完整路径" or file_entry.get() == "Please enter the full path of the Excel file to be analyzed":
        file_entry.delete(0, tk.END)
        file_entry.configure(style="TEntry")  # 恢复默认样式


def on_focusout(event):
    """当用户离开输入框时，如果没有输入内容，恢复提示信息"""
    if file_entry.get() == "":
        if current_language == 'zh':
            file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
            file_entry.configure(style="Gray.TEntry")
        else:
            file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
            file_entry.configure(style="Gray.TEntry")


# 创建主窗口
root = ttk.Window(themename="flatly")
root.title(LANGUAGES[current_language]['title'])

# 获取屏幕的宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 设置窗口的宽度和高度
window_width = 500
window_height = 300

# 计算窗口应该放置的位置
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2

# 设置窗口的位置和大小
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 创建自定义样式
style = ttk.Style()
style.configure("Gray.TEntry", foreground="gray")

# 创建文件选择按钮
select_button = ttk.Button(root, text=LANGUAGES[current_language]['select_button'], command=select_file,
                           bootstyle=PRIMARY)
select_button.pack(pady=10)

# 创建文件路径输入框
file_entry = ttk.Entry(root, width=50, style="Gray.TEntry")
if current_language == 'zh':
    file_entry.insert(0, "请输入待分析 Excel 文件的完整路径")
else:
    file_entry.insert(0, "Please enter the full path of the Excel file to be analyzed")
file_entry.pack(pady=5)
file_entry.bind("<FocusIn>", on_entry_click)
file_entry.bind("<FocusOut>", on_focusout)

# 创建分析按钮
analyze_button = ttk.Button(root, text=LANGUAGES[current_language]['analyze_button'], command=analyze_file,
                            bootstyle=SUCCESS)
analyze_button.pack(pady=10)

# 创建语言切换标签
language_label = ttk.Label(root, text=LANGUAGES[current_language]['switch_language'], cursor="hand2")
language_label.pack(pady=10)
language_label.bind("<Button-1>", lambda event: switch_language())

# 创建结果显示标签
result_label = ttk.Label(root, text="", justify=tk.LEFT)
result_label.pack(pady=10)

# 运行主循环
root.mainloop()